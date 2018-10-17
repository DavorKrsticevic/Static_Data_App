from tkinter import *
from tkinter import ttk
from tkinter.filedialog import askopenfilename
import os
import openpyxl
from openpyxl import load_workbook
import xlsxwriter
import pyexcel as p
import time
import pyexcel



exceptions_file_path = ''
data_file_path = ''
company_code = ''


class StaticData:
        

    def __init__(self, master):  
        
        master.title('Static Data App © davork')
        master.configure(background = '#ffbb33')

        self.style = ttk.Style()
        self.style.configure('TFrame', background = '#ffbb33')
        self.style.configure('TButton', background = '#ffbb33', font = ('Arial', 10))
        self.style.configure('TLabel', background = '#ffbb33', font = ('Arial', 10))
        self.style.configure('Header.TLabel', font = ('Arial', 18, 'bold'))      
        

        self.frame_header = ttk.Frame(master)
        self.frame_header.pack()
        
        
        ttk.Label(self.frame_header, text = 'Static Data App!',style = 'Header.TLabel').grid(row = 0, column = 1)
        ttk.Label(self.frame_header, wraplength = 400,
                  text = ("Just click on the buttons one by one and follow the logic")).grid(row = 1, column = 1)
       


        self.frame_content = ttk.Frame(master)
        self.frame_content.pack()       

        ttk.Button(self.frame_content, text = 'Ok , let us start with loading your exceptions file first!', command=self.load_ex_file).grid(row = 6, column = 0, columnspan = 2 , padx = 0, pady = 20 )
        ttk.Button(self.frame_content, text = 'Now , let us load your data file as well!', command=self.load_data_file).grid(row = 7, column = 0, columnspan = 2 , padx = 0, pady = 20 )
        ttk.Button(self.frame_content, text = 'Let\'s go!', command=self.lets_go).grid(row = 8, column = 0, columnspan = 2, padx = 0, pady = 50 )
        
        
        
        
        
        

    # function to load paths for data and exception file
        

    def load_ex_file(self):
        global exceptions_file_path 
        exceptions_file_path = askopenfilename(filetypes=(("Exceptions File", "*.xls"),
                                           ("HTML files", "*.html;*.htm"),
                                           ("All files", "*.*") ))
        print(exceptions_file_path)
        
        
    def load_data_file(self):
        global data_file_path
        data_file_path = askopenfilename(filetypes=(("EMT, EPT or AAA", "*.xlsx"),
                                           ("HTML files", "*.html;*.htm"),
                                           ("All files", "*.*") ))  
    
        

    def lets_go(self):

        
        widget = ttk.Label(self.frame_content, text = 'Your Files Are READY!!!',background = '#ffff80', font = ('Arial', 15, 'bold')).grid(row = 8, column = 0, columnspan = 2, padx = 0, pady = 50)
        
               
        #widget.config(height=3, width=20)

        if 'EMT' in exceptions_file_path:
            EMT(self) 
        if 'EPT' in exceptions_file_path:
            EPT(self)
        if 'AAA' in exceptions_file_path:
            AAA(self) 

        
    

def EMT(self):
    # Convert exceptions xls file into xlsx

    p.save_book_as(file_name= exceptions_file_path,
        dest_file_name='exceptions.xlsx')


    # Load xlsx files as worksheets

    wb = openpyxl.load_workbook(data_file_path)
    wb2 = load_workbook('exceptions.xlsx')


    # Read the proper sheet from the files 

    sheet_number_datafile = wb.sheetnames
    sheet_number_exceptionsfile = wb2.sheetnames

    sheet = wb[sheet_number_datafile[0]] 
    sheet2 = wb2[sheet_number_exceptionsfile[1]]


    # Retrieve cell value

    rows_exceptions_file = []
    save = ['S']
    identifier = ['ISIN']
    ISIN, seen , duplicate_index = [], set(), []
    name = []
    currency = []
    character_F = ['F']
    rows_string = []
    duplicate_index = []
    company_code =  exceptions_file_path[-41 : - 38]


    # read exceptions xlsx file and save data into a variable 
 
    for col in sheet2.iter_cols(min_row=1, min_col= 1 ,max_col=1):
        for cell in col:
            if cell.value == "Identifier resolution":
                rows_exceptions_file.append(cell.row)

    for i in rows_exceptions_file:
        rows_string.append(sheet2.cell(row=i , column=4).value)



    # Save values in a variable


    rows = [int(item) for item in rows_string]


    for i in rows:
        ISIN.append(sheet.cell(row=i, column=1).value)

    for i in rows:
        name.append(sheet.cell(row=i, column=3).value)

    for i in rows:
        currency.append(sheet.cell(row=i, column=4).value)
    

    # remove the duplicates

    for idx, item in enumerate(ISIN):
        if item not in seen:
            seen.add(item)
        else:
            duplicate_index.append(idx)
    
    print(duplicate_index)

    duplicate_index = list(map(int, duplicate_index))


    for i in sorted(duplicate_index, reverse = True):
        del ISIN[i]
        del name[i]
        del currency[i]


    # itarate over number of ISIN's

    number_of_rows = len(ISIN)

    # Get todays date

    timestr = time.strftime("%Y%m")



    # Create an new Excel file and add a worksheet for Funddata
    
    
    workbook = xlsxwriter.Workbook('StaticData\Funddata_'+(company_code)+'_%s.xlsx' %timestr)
    worksheet = workbook.add_worksheet()
    

    # Write some numbers, with row/column notation.


    worksheet.write_column(0, 1, save * number_of_rows)
    worksheet.write_column(0, 2, identifier * number_of_rows)
    worksheet.write_column(0, 3, ISIN)
    worksheet.write_column(0, 4, name)


    # Set coulmns width

    worksheet.set_column('A:A', 15)
    worksheet.set_column('B:C', 4)
    worksheet.set_column('D:D', 13)
    worksheet.set_column('E:E', 55)

    # delete exceptions.xlsx

    os.remove("exceptions.xlsx")


    workbook.close()



    # Create an new Excel file and add a worksheet for Shareclassdata

    workbook = xlsxwriter.Workbook('StaticData\Shareclassdata_'+(company_code)+'_%s.xlsx' %timestr)
    worksheet = workbook.add_worksheet()


    # Write some numbers, with row/column notation.


    worksheet.write_column(0, 1, save * number_of_rows)
    worksheet.write_column(0, 2, identifier * number_of_rows)
    worksheet.write_column(0, 3, ISIN)
    worksheet.write_column(0, 4, identifier * number_of_rows)
    worksheet.write_column(0, 5, ISIN)
    worksheet.write_column(0, 6, name)
    worksheet.write_column(0, 7, currency)
    worksheet.write_column(0, 8, character_F * number_of_rows)

    # Set coulmns width

    worksheet.set_column('A:A', 15)
    worksheet.set_column('B:C', 4)
    worksheet.set_column('D:D', 13)
    worksheet.set_column('E:E', 4)
    worksheet.set_column('F:F', 13)
    worksheet.set_column('G:G', 55)
    worksheet.set_column('H:I', 4)


    workbook.close()
    


def EPT(self):
    # Convert exceptions xls file into xlsx

    p.save_book_as(file_name= exceptions_file_path,
        dest_file_name='./exceptions.xlsx')


    # Load xlsx files as worksheets

    wb = openpyxl.load_workbook(data_file_path)
    wb2 = load_workbook('./exceptions.xlsx')


    # Read the proper sheet from the files 

    sheet_number_datafile = wb.sheetnames
    sheet_number_exceptionsfile = wb2.sheetnames

    sheet = wb[sheet_number_datafile[0]] 
    sheet2 = wb2[sheet_number_exceptionsfile[1]]


    # Retrieve cell value

    rows_exceptions_file = []
    save = ['S']
    identifier = ['ISIN']
    ISIN, seen , duplicate_index = [], set(), []
    name = []
    currency = []
    character_F = ['F']
    rows_string = []
    duplicate_index = []
    company_code =  exceptions_file_path[-41 : - 38]


    # read exceptions xlsx file and save data into a variable 
 
    for col in sheet2.iter_cols(min_row=1, min_col= 1 ,max_col=1):
        for cell in col:
            if cell.value == "Identifier resolution":
                rows_exceptions_file.append(cell.row)

    for i in rows_exceptions_file:
        rows_string.append(sheet2.cell(row=i , column=4).value)



    # Save values in a variable


    rows = [int(item) for item in rows_string]


    for i in rows:
        ISIN.append(sheet.cell(row=i, column=3).value)

    for i in rows:
        name.append(sheet.cell(row=i, column=5).value)

    for i in rows:
        currency.append(sheet.cell(row=i, column=6).value)
    

    # remove the duplicates

    for idx, item in enumerate(ISIN):
        if item not in seen:
            seen.add(item)
        else:
            duplicate_index.append(idx)
    
    print(duplicate_index)

    duplicate_index = list(map(int, duplicate_index))


    for i in sorted(duplicate_index, reverse = True):
        del ISIN[i]
        del name[i]
        del currency[i]


    # itarate over number of ISIN's

    number_of_rows = len(ISIN)

    # Get todays date

    timestr = time.strftime("%Y%m")



    # Create an new Excel file and add a worksheet for Funddata

    workbook = xlsxwriter.Workbook('StaticData\Funddata_'+(company_code)+'_%s.xlsx' %timestr)
    worksheet = workbook.add_worksheet()

    # Write some numbers, with row/column notation.


    worksheet.write_column(0, 1, save * number_of_rows)
    worksheet.write_column(0, 2, identifier * number_of_rows)
    worksheet.write_column(0, 3, ISIN)
    worksheet.write_column(0, 4, name)


    # Set coulmns width

    worksheet.set_column('A:A', 15)
    worksheet.set_column('B:C', 4)
    worksheet.set_column('D:D', 13)
    worksheet.set_column('E:E', 55)


    workbook.close()



    # Create an new Excel file and add a worksheet for Shareclassdata

    workbook = xlsxwriter.Workbook('StaticData\Shareclassdata_'+(company_code)+'_%s.xlsx' %timestr)
    worksheet = workbook.add_worksheet()


    # Write some numbers, with row/column notation.


    worksheet.write_column(0, 1, save * number_of_rows)
    worksheet.write_column(0, 2, identifier * number_of_rows)
    worksheet.write_column(0, 3, ISIN)
    worksheet.write_column(0, 4, identifier * number_of_rows)
    worksheet.write_column(0, 5, ISIN)
    worksheet.write_column(0, 6, name)
    worksheet.write_column(0, 7, currency)
    worksheet.write_column(0, 8, character_F * number_of_rows)

    # Set coulmns width

    worksheet.set_column('A:A', 15)
    worksheet.set_column('B:C', 4)
    worksheet.set_column('D:D', 13)
    worksheet.set_column('E:E', 4)
    worksheet.set_column('F:F', 13)
    worksheet.set_column('G:G', 55)
    worksheet.set_column('H:I', 4)

    # delete exceptions.xlsx

    os.remove("./exceptions.xlsx")


    workbook.close()   
    

def AAA(self):
    # Convert exceptions xls file into xlsx

    p.save_book_as(file_name= exceptions_file_path,
        dest_file_name='./exceptions.xlsx')


    # Load xlsx files as worksheets

    wb = openpyxl.load_workbook(data_file_path)
    wb2 = load_workbook('./exceptions.xlsx')


    # Read the proper sheet from the files 

    sheet_number_datafile = wb.sheetnames
    sheet_number_exceptionsfile = wb2.sheetnames

    sheet = wb[sheet_number_datafile[0]] 
    sheet2 = wb2[sheet_number_exceptionsfile[1]]


    # Retrieve cell value

    rows_exceptions_file = []
    save = ['S']
    identifier = ['ISIN']
    ISIN, seen , duplicate_index = [], set(), []
    name = []
    currency = []
    character_F = ['F']
    rows_string = []
    duplicate_index = []
    company_code =  exceptions_file_path[-39 : - 36]


    # read exceptions xlsx file and save data into a variable 
 
    for col in sheet2.iter_cols(min_row=1, min_col= 1 ,max_col=1):
        for cell in col:
            if cell.value == "Identifier resolution":
                rows_exceptions_file.append(cell.row)

    for i in rows_exceptions_file:
        rows_string.append(sheet2.cell(row=i , column=4).value)



    # Save values in a variable


    rows = [int(item) for item in rows_string]


    for i in rows:
        ISIN.append(sheet.cell(row=i, column=1).value)

    for i in rows:
        name.append(sheet.cell(row=i, column=3).value)

    for i in rows:
        currency.append(sheet.cell(row=i, column=4).value)
    

    # remove the duplicates

    for idx, item in enumerate(ISIN):
        if item not in seen:
            seen.add(item)
        else:
            duplicate_index.append(idx)
    
    print(duplicate_index)

    duplicate_index = list(map(int, duplicate_index))


    for i in sorted(duplicate_index, reverse = True):
        del ISIN[i]
        del name[i]
        del currency[i]


    # itarate over number of ISIN's

    number_of_rows = len(ISIN)

    # Get todays date

    timestr = time.strftime("%Y%m")



    # Create an new Excel file and add a worksheet for Funddata

    workbook = xlsxwriter.Workbook('StaticData\Funddata_'+(company_code)+'_%s.xlsx' %timestr)
    worksheet = workbook.add_worksheet()

    # Write some numbers, with row/column notation.


    worksheet.write_column(0, 1, save * number_of_rows)
    worksheet.write_column(0, 2, identifier * number_of_rows)
    worksheet.write_column(0, 3, ISIN)
    worksheet.write_column(0, 4, name)


    # Set coulmns width

    worksheet.set_column('A:A', 15)
    worksheet.set_column('B:C', 4)
    worksheet.set_column('D:D', 13)
    worksheet.set_column('E:E', 55)


    workbook.close()



    # Create an new Excel file and add a worksheet for Shareclassdata

    workbook = xlsxwriter.Workbook('StaticData\Shareclassdata_'+(company_code)+'_%s.xlsx' %timestr)
    worksheet = workbook.add_worksheet()


    # Write some numbers, with row/column notation.


    worksheet.write_column(0, 1, save * number_of_rows)
    worksheet.write_column(0, 2, identifier * number_of_rows)
    worksheet.write_column(0, 3, ISIN)
    worksheet.write_column(0, 4, identifier * number_of_rows)
    worksheet.write_column(0, 5, ISIN)
    worksheet.write_column(0, 6, name)
    worksheet.write_column(0, 7, currency)
    worksheet.write_column(0, 8, character_F * number_of_rows)

    # Set coulmns width

    worksheet.set_column('A:A', 15)
    worksheet.set_column('B:C', 4)
    worksheet.set_column('D:D', 13)
    worksheet.set_column('E:E', 4)
    worksheet.set_column('F:F', 13)
    worksheet.set_column('G:G', 55)
    worksheet.set_column('H:I', 4)

    # delete exceptions.xlsx

    os.remove("./exceptions.xlsx")

    workbook.close()


    

root = Tk()
feedback = StaticData(root)
root.geometry("550x335+650+250")
root.resizable(False, False)            
root.mainloop()

    
